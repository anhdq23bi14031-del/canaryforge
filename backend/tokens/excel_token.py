"""
Excel Canary Token — creates an .xlsx file with convincing LLM-generated
tabular decoy data and an embedded tracking mechanism.

Caveat (same as PDF): Excel does not reliably auto-fetch remote content on
open across all versions/platforms (and many orgs block external content
by default for security). A visible "source" hyperlink cell is included as
the primary, reliable trigger mechanism — clicking it (or automated link-
following tools scanning the file) fires the canary.
"""
import io
import secrets
from typing import Optional

from backend.config import settings
from backend.intelligence.llm_engine import get_engine, TabularDecoyContent


async def generate_excel_token(
    token_id: str,
    name: str,
    content_type: str = "financial",
    company_hint: Optional[str] = None,
    use_llm: bool = True,
) -> dict:
    """
    Generate a canary Excel spreadsheet with convincing LLM-written tabular content.
    Primary trigger: a hyperlink cell in the sheet.
    """
    slug = secrets.token_urlsafe(16)
    click_url = f"{settings.BASE_URL}/sheets/{slug}"

    if use_llm and settings.ANTHROPIC_API_KEY:
        engine = get_engine()
        content: TabularDecoyContent = await engine.generate_tabular_content(
            content_type=content_type,
            company_hint=company_hint,
        )
        filename = engine.suggest_filename(content_type, content.keywords, extension="xlsx")
    else:
        content = _fallback_content(content_type)
        filename = f"confidential_{content_type}_data.xlsx"

    xlsx_bytes = _build_xlsx(content, click_url)

    return {
        "token_value": click_url,
        "slug": slug,
        "filename": filename,
        "doc_bytes": xlsx_bytes,
        "metadata": {
            "name": name,
            "type": "excel",
            "slug": slug,
            "tracking_url": click_url,
            "content_type": content_type,
            "doc_title": content.title,
            "doc_summary": content.summary,
            "filename": filename,
            "instructions": (
                f"Drop '{filename}' in a file share or email attachment. "
                f"The sheet contains a 'source' hyperlink cell that triggers the canary when clicked. "
                f"Note: Excel does not reliably auto-fetch remote content on open, "
                f"so the hyperlink cell is the primary, reliable trigger."
            ),
        },
    }


def _build_xlsx(content: TabularDecoyContent, click_url: str) -> bytes:
    """
    Build an .xlsx file with decoy tabular content using openpyxl.
    Includes a styled header row, data rows, and a hyperlink "source" cell.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = content.sheet_name or "Sheet1"

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(content.headers), 1))
    title_cell = ws.cell(row=1, column=1, value=content.title)
    title_cell.font = Font(size=14, bold=True, color="1A1A2E")
    ws.row_dimensions[1].height = 24

    # Header row (row 3, leave row 2 blank for spacing)
    header_row = 3
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, header in enumerate(content.headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left")

    # Data rows
    for row_offset, row_data in enumerate(content.rows, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=header_row + row_offset, column=col_idx, value=value)

    # Auto-size columns roughly
    for col_idx in range(1, len(content.headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            [len(str(content.headers[col_idx - 1]))]
            + [len(str(r[col_idx - 1])) for r in content.rows if col_idx - 1 < len(r)]
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)

    # "Source" hyperlink row — the reliable trigger mechanism
    source_row = header_row + len(content.rows) + 2
    source_cell = ws.cell(row=source_row, column=1, value="Source / last updated reference")
    source_cell.font = Font(size=8, color="999999", italic=True)
    link_cell = ws.cell(row=source_row + 1, column=1, value=click_url)
    link_cell.hyperlink = click_url
    link_cell.font = Font(size=8, color="999999", underline="single")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fallback_content(content_type: str) -> TabularDecoyContent:
    return TabularDecoyContent(
        title=f"Confidential {content_type.title()} Data — Internal Use Only",
        sheet_name="Data",
        headers=["Item", "Value", "Notes"],
        rows=[
            ["CONFIDENTIAL", "DO NOT DISTRIBUTE", "Internal use only"],
            ["Contact", "document owner", "For questions"],
        ],
        summary=f"Internal {content_type} spreadsheet",
        keywords=[content_type, "confidential", "internal"],
    )
